"""Day 7. Excel writer for BRIEF.

Six sheets:
  1. Summary       - headline KPIs, exec summary, embedded tone curve PNG.
  2. Guidance      - one row per forward guidance line item.
  3. Themes        - themes with weight, sentiment, key quotes.
  4. QA            - analyst-by-analyst Q&A grid.
  5. Quotes        - phrase hits + number claims, with speaker + context.
  6. Inputs        - parsed metadata + the raw transcript echoed for audit.
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from analysis import CallAnalysis
from brief_chart import render_speaker_breakdown_png, render_tone_curve_png
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY_FILL    = PatternFill("solid", fgColor="0A1628")
ORANGE_FILL  = PatternFill("solid", fgColor="FF6B00")
CYAN_FILL    = PatternFill("solid", fgColor="06B6D4")
PANEL_FILL   = PatternFill("solid", fgColor="F4F1EA")
RED_FILL     = PatternFill("solid", fgColor="FEE2E2")
AMBER_FILL   = PatternFill("solid", fgColor="FEF3C7")
GREEN_FILL   = PatternFill("solid", fgColor="D1FAE5")

HEADER_FONT = Font(bold=True, color="F4F1EA", size=11)
LABEL_FONT  = Font(bold=True, color="0A1628", size=10)
TITLE_FONT  = Font(bold=True, color="0A1628", size=14)


def write_workbook(call: CallAnalysis, out_path: Path) -> Path:
    wb = Workbook()
    _write_summary(wb.active, call)
    wb.active.title = "Summary"
    _write_guidance(wb.create_sheet("Guidance"), call)
    _write_themes(wb.create_sheet("Themes"), call)
    _write_qa(wb.create_sheet("QA"), call)
    _write_quotes(wb.create_sheet("Quotes"), call)
    _write_inputs(wb.create_sheet("Inputs"), call)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_summary(ws, call: CallAnalysis) -> None:
    md = call.transcript.metadata
    h = call.headline
    ws["A1"] = f"BRIEF | {md.company} | {md.fiscal_period}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    rows = [
        ("Ticker", md.ticker),
        ("Call date", md.call_date),
        ("Word count", md.word_count),
        ("Approx minutes", h.minutes),
        ("Distinct analysts", h.analyst_count),
        ("Overall tone", h.overall_tone),
        ("Confidence score", round(h.confidence_score, 3)),
        ("Hedge phrases", h.hedge_count),
        ("Certainty phrases", h.certainty_count),
        ("Deflection phrases", h.deflection_count),
        ("Quantitative claims", h.quantitative_claims),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    if call.exec_summary:
        ws.cell(row=15, column=1, value="Verdict").font = LABEL_FONT
        ws.cell(row=15, column=2, value=call.exec_summary.headline)
        ws.cell(row=16, column=1, value="Bull case").font = LABEL_FONT
        ws.cell(row=16, column=2, value=call.exec_summary.bull_case)
        ws.cell(row=17, column=1, value="Bear case").font = LABEL_FONT
        ws.cell(row=17, column=2, value=call.exec_summary.bear_case)
        for i, a in enumerate(call.exec_summary.actions[:5]):
            ws.cell(row=18 + i, column=1,
                    value="Action" if i == 0 else "").font = LABEL_FONT
            ws.cell(row=18 + i, column=2, value=a)
    for r in range(15, 25):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["B"].width = 80

    # Embed tone curve + speaker breakdown PNGs
    try:
        png = render_tone_curve_png(call.tone_curve)
        img = XLImage(BytesIO(png))
        img.width, img.height = 720, 260
        ws.add_image(img, "D3")
    except Exception:
        pass
    try:
        png = render_speaker_breakdown_png(h.role_word_share)
        img = XLImage(BytesIO(png))
        img.width, img.height = 360, 280
        ws.add_image(img, "D18")
    except Exception:
        pass


def _write_guidance(ws, call: CallAnalysis) -> None:
    headers = ["Metric", "Period", "Direction", "Range low", "Range high",
               "Unit", "Quote", "Prior"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for i, g in enumerate(call.guidance, start=2):
        ws.cell(row=i, column=1, value=g.metric)
        ws.cell(row=i, column=2, value=g.period)
        cell = ws.cell(row=i, column=3, value=g.direction)
        if g.direction in ("up", "raised"):
            cell.fill = GREEN_FILL
        elif g.direction in ("down", "lowered", "withdrawn"):
            cell.fill = RED_FILL
        elif g.direction == "flat":
            cell.fill = AMBER_FILL
        ws.cell(row=i, column=4, value=g.range_low)
        ws.cell(row=i, column=5, value=g.range_high)
        ws.cell(row=i, column=6, value=g.unit)
        ws.cell(row=i, column=7, value=g.quote).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=8, value=g.prior).alignment = Alignment(wrap_text=True)
    for col, w in enumerate([18, 14, 14, 12, 12, 14, 60, 40], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_themes(ws, call: CallAnalysis) -> None:
    headers = ["Theme", "Weight", "Sentiment", "Key quotes"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for i, t in enumerate(call.themes, start=2):
        ws.cell(row=i, column=1, value=t.name)
        ws.cell(row=i, column=2, value=t.weight)
        cell = ws.cell(row=i, column=3, value=t.sentiment)
        if t.sentiment == "bullish":
            cell.fill = GREEN_FILL
        elif t.sentiment == "bearish":
            cell.fill = RED_FILL
        else:
            cell.fill = AMBER_FILL
        ws.cell(row=i, column=4, value="\n\n".join(t.key_quotes)).alignment = Alignment(wrap_text=True)
    for col, w in enumerate([26, 10, 14, 90], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_qa(ws, call: CallAnalysis) -> None:
    headers = ["Analyst", "Firm", "Question", "Answer", "Tension", "Clarity"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for i, q in enumerate(call.qa, start=2):
        ws.cell(row=i, column=1, value=q.analyst_name)
        ws.cell(row=i, column=2, value=q.analyst_firm)
        ws.cell(row=i, column=3, value=q.question_summary).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=4, value=q.answer_summary).alignment = Alignment(wrap_text=True)
        cell = ws.cell(row=i, column=5, value=q.tension)
        if q.tension == "hostile":
            cell.fill = RED_FILL
        elif q.tension == "probing":
            cell.fill = AMBER_FILL
        else:
            cell.fill = GREEN_FILL
        cell2 = ws.cell(row=i, column=6, value=q.management_clarity)
        if q.management_clarity == "clear":
            cell2.fill = GREEN_FILL
        elif q.management_clarity == "hedged":
            cell2.fill = AMBER_FILL
        else:
            cell2.fill = RED_FILL
    for col, w in enumerate([22, 22, 50, 60, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_quotes(ws, call: CallAnalysis) -> None:
    ws["A1"] = "Tone signal hits"
    ws["A1"].font = TITLE_FONT
    headers = ["Bucket", "Phrase", "Speaker", "Role", "Minute", "Context"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=name)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
    row = 3
    for h in call.phrase_hits[:200]:
        cell = ws.cell(row=row, column=1, value=h.bucket)
        if h.bucket == "certainty":
            cell.fill = GREEN_FILL
        elif h.bucket == "deflection":
            cell.fill = RED_FILL
        else:
            cell.fill = AMBER_FILL
        ws.cell(row=row, column=2, value=h.phrase)
        ws.cell(row=row, column=3, value=h.speaker)
        ws.cell(row=row, column=4, value=h.role)
        ws.cell(row=row, column=5, value=h.minute)
        ws.cell(row=row, column=6, value=h.context).alignment = Alignment(wrap_text=True)
        row += 1
    row += 2
    ws.cell(row=row, column=1, value="Quantitative claims").font = TITLE_FONT
    row += 1
    for col, name in enumerate(["Raw", "Kind", "Value low", "Value high",
                                "Unit", "Speaker", "Role", "Context"], start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.fill = NAVY_FILL
        c.font = HEADER_FONT
    row += 1
    for n in call.number_claims[:200]:
        ws.cell(row=row, column=1, value=n.raw)
        ws.cell(row=row, column=2, value=n.kind)
        ws.cell(row=row, column=3, value=n.value_low)
        ws.cell(row=row, column=4, value=n.value_high)
        ws.cell(row=row, column=5, value=n.unit)
        ws.cell(row=row, column=6, value=n.speaker)
        ws.cell(row=row, column=7, value=n.role)
        ws.cell(row=row, column=8, value=n.context).alignment = Alignment(wrap_text=True)
        row += 1
    for col, w in enumerate([14, 26, 22, 14, 8, 60], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_inputs(ws, call: CallAnalysis) -> None:
    md = call.transcript.metadata
    ws["A1"] = "Inputs (echoed for audit)"
    ws["A1"].font = TITLE_FONT
    pairs = [
        ("Company", md.company),
        ("Ticker", md.ticker),
        ("Fiscal period", md.fiscal_period),
        ("Call date", md.call_date),
        ("Word count", md.word_count),
        ("Source filename", md.source_filename),
    ]
    for i, (k, v) in enumerate(pairs, start=3):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
